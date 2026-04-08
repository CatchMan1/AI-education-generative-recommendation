"""
majors_encode.py
────────────────────────────────────────────────────────────────
为三级专业层级结构生成语义中心向量，并保存到 HDF5 文件。

数据来源：majors_hierarchy.xlsx
  Level1 sheet：大类（大类编码 | 大类 | 大类简介及培养目标 | 大类通识课程）
  Level2 sheet：专业（专业编码 | 专业名 | 专业简介及培养目标 | 专业核心课程）
  Level3 sheet：学科大类（学类编码 | 学类名 | 学类简介及培养目标）

输出（HDF5）：
  /Volumes/study/AI_edu/bert/level1_embs.h5
  /Volumes/study/AI_edu/bert/level2_embs.h5
  /Volumes/study/AI_edu/bert/level3_embs.h5
  包含 datasets: codes, names, embs, meta
"""

import argparse
import json
import os

# ── 允许在线加载 HuggingFace 模型（方案 A）──
# 删除文件中强制设置为离线的环境变量，以便在本机无缓存时可以从 Hugging Face 下载模型。
# 运行时请保证网络可用，私有模型则需先登录 huggingface-cli。
os.environ.setdefault("HF_HUB_DISABLE_TELEMETRY", "1")
os.environ.setdefault("HF_HUB_DISABLE_IMPLICIT_TOKEN", "1")
os.environ.setdefault("TRANSFORMERS_VERBOSITY", "error")
os.environ.setdefault("DISABLE_TELEMETRY", "1")
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")  # 避免 tokenizer 并行警告

import warnings
warnings.filterwarnings("ignore", message=".*Discussions are disabled.*")
warnings.filterwarnings("ignore", message=".*unauthenticated requests.*")
warnings.filterwarnings("ignore", message=".*You are sending unauthenticated.*")

import h5py
import numpy as np
import torch
from openpyxl import load_workbook
from tqdm import tqdm
from transformers import AutoModel, AutoTokenizer, logging as hf_logging

hf_logging.set_verbosity_error()


# ──────────────────────────────────────────────
# 1. 读取 Excel
# ──────────────────────────────────────────────

def load_level1(path: str) -> list[tuple[str, str, str]]:
    """
    返回 list of (code, name, text)
    text = 大类简介及培养目标 + 大类通识课程
    """
    wb = load_workbook(path, read_only=True)
    ws = wb["Level1"]
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        code, name, intro, courses = row[0], row[1], row[2], row[3]
        if code is None:
            continue
        parts = []
        if intro:
            parts.append(str(intro))
        if courses:
            parts.append("大类通识课程：" + str(courses))
        rows.append((str(code), str(name), " ".join(parts)))
    wb.close()
    return rows


def load_level2(path: str) -> list[tuple[str, str, str]]:
    """
    返回 list of (code, name, text)
    text = 专业简介及培养目标 + 专业核心课程
    同一 code 可能对应多个分流专业，各自独立成行。
    """
    wb = load_workbook(path, read_only=True)
    ws = wb["Level2"]
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        code, name, intro, courses = row[0], row[1], row[2], row[3]
        if code is None:
            continue
        parts = []
        if intro:
            parts.append(str(intro))
        if courses:
            parts.append("专业核心课程：" + str(courses))
        rows.append((str(code), str(name), " ".join(parts)))
    wb.close()
    return rows


def load_level3(path: str) -> list[tuple[str, str, str]]:
    """
    返回 list of (code, name, text)
    直接使用 Level3 sheet 第三列"学类简介及培养目标"作为编码文本。
    若简介为空则退回到学科名称本身（兜底）。
    """
    wb = load_workbook(path, read_only=True)
    ws = wb["Level3"]
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        code, name = row[0], row[1]
        intro = row[2] if len(row) > 2 else None
        if code is None:
            continue
        text = str(intro).strip() if intro else str(name)
        rows.append((str(code), str(name), text))
    wb.close()
    return rows


# ──────────────────────────────────────────────
# 2. BERT 编码（去 [CLS]，均值池化）
# ──────────────────────────────────────────────

@torch.no_grad()
def encode_texts(
    texts: list[str],
    tokenizer,
    model,
    device: torch.device,
    batch_size: int,
    max_length: int,
) -> np.ndarray:
    """
    将文本列表编码为向量矩阵。
    池化策略：去掉 [CLS] token（index 0），对剩余 token 做 masked 均值池化。
    返回 np.ndarray, shape (N, D), dtype float32
    """
    model.eval()
    embs = []
    for start in tqdm(range(0, len(texts), batch_size), desc="Encoding"):
        batch = texts[start: start + batch_size]
        encoded = tokenizer(
            batch,
            padding=True,
            truncation=True,
            max_length=max_length,
            return_tensors="pt",
        ).to(device)

        outputs = model(**encoded)
        hidden = outputs.last_hidden_state          # (B, L, D)
        mask   = encoded["attention_mask"]          # (B, L)
        masked = hidden * mask.unsqueeze(-1)        # (B, L, D)
        # 去掉 index=0 的 [CLS]，对 index 1: 做均值；clamp 防止全 padding 的除零
        mean_emb = (
            masked[:, 1:, :].sum(dim=1)
            / mask[:, 1:].sum(dim=-1, keepdim=True).clamp(min=1e-9)
        )
        embs.append(mean_emb.detach().cpu())

    return torch.cat(embs, dim=0).numpy().astype(np.float32)


# ──────────────────────────────────────────────
# 3. 保存到 HDF5
# ──────────────────────────────────────────────

def save_h5(
    out_path: str,
    codes: list[str],
    names: list[str],
    embs: np.ndarray,
    meta: dict,
) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with h5py.File(out_path, "w") as f:
        dt = h5py.string_dtype(encoding="utf-8")
        f.create_dataset("codes", data=np.array(codes, dtype=object), dtype=dt)
        f.create_dataset("names", data=np.array(names, dtype=object), dtype=dt)
        f.create_dataset("embs",  data=embs, compression="gzip")
        meta_bytes = json.dumps(meta, ensure_ascii=False).encode("utf-8")
        f.create_dataset("meta",  data=np.bytes_(meta_bytes))
    print(f"  ✓ saved  {out_path}  shape={embs.shape}")


# ──────────────────────────────────────────────
# 4. 主流程
# ──────────────────────────────────────────────

def encode(params: dict) -> None:
    device = (
        torch.device(params["device"])
        if params.get("device")
        else torch.device("cuda" if torch.cuda.is_available() else "cpu")
    )
    print(f"Using device: {device}")

    xlsx_path    = params["xlsx_path"]
    encode_model = params["encode_model"]
    batch_size   = params["encode_batch_size"]
    max_length   = params["encode_max_len"]
    out_dir      = params["output_dir"]

    # ── 读取数据 ──────────────────────────────────
    print("Loading Excel …")
    level1_rows = load_level1(xlsx_path)
    level2_rows = load_level2(xlsx_path)
    level3_rows = load_level3(xlsx_path)
    print(f"  Level1 (大类专业): {len(level1_rows)} rows")
    print(f"  Level2 (分流专业): {len(level2_rows)} rows")
    print(f"  Level3 (学科大类): {len(level3_rows)} rows")

    # ── 加载模型（允许在线下载/更新）──────────────────
    print(f"\nLoading model: {encode_model} …")
    tokenizer = AutoTokenizer.from_pretrained(encode_model)
    model = AutoModel.from_pretrained(
        encode_model,
        use_safetensors=False,   # 避免转换检查
        local_files_only=False,  # 允许在线下载模型（若本地缓存不存在）
    ).to(device)
    print("  Model loaded.")

    meta_base = {
        "model":      encode_model,
        "max_length": max_length,
        "pooling":    "mean_no_cls",
    }

    # ── Level 1 ───────────────────────────────────
    print("\n[Level1] Encoding 大类专业 …")
    l1_codes = [r[0] for r in level1_rows]
    l1_names = [r[1] for r in level1_rows]
    l1_texts = [r[2] for r in level1_rows]
    l1_embs  = encode_texts(l1_texts, tokenizer, model, device, batch_size, max_length)
    save_h5(
        os.path.join(out_dir, "level1_embs.h5"),
        l1_codes, l1_names, l1_embs,
        {**meta_base, "level": "Level1_大类专业", "dim": int(l1_embs.shape[1])},
    )

    # ── Level 2 ───────────────────────────────────
    print("\n[Level2] Encoding 分流专业 …")
    l2_codes = [r[0] for r in level2_rows]
    l2_names = [r[1] for r in level2_rows]
    l2_texts = [r[2] for r in level2_rows]
    l2_embs  = encode_texts(l2_texts, tokenizer, model, device, batch_size, max_length)
    save_h5(
        os.path.join(out_dir, "level2_embs.h5"),
        l2_codes, l2_names, l2_embs,
        {**meta_base, "level": "Level2_分流专业", "dim": int(l2_embs.shape[1])},
    )

    # ── Level 3 ───────────────────────────────────
    print("\n[Level3] Encoding 学科大类 …")
    l3_codes = [r[0] for r in level3_rows]
    l3_names = [r[1] for r in level3_rows]
    l3_texts = [r[2] for r in level3_rows]
    l3_embs  = encode_texts(l3_texts, tokenizer, model, device, batch_size, max_length)
    save_h5(
        os.path.join(out_dir, "level3_embs.h5"),
        l3_codes, l3_names, l3_embs,
        {**meta_base, "level": "Level3_学科大类", "dim": int(l3_embs.shape[1])},
    )

    print("\nAll done.")


# ──────────────────────────────────────────────
# 5. CLI / Jupyter 入口
# ──────────────────────────────────────────────

def parse_args() -> dict:
    p = argparse.ArgumentParser(description="Build major-hierarchy semantic embeddings")
    p.add_argument("--xlsx_path",         default="majors_hierarchy.xlsx")
    p.add_argument("--encode_model",      default="hfl/chinese-roberta-wwm-ext",
                   help="HuggingFace model name or local path")
    p.add_argument("--encode_batch_size", type=int, default=32)
    p.add_argument("--encode_max_len",    type=int, default=512)
    # 修改默认输出路径为用户指定的目录
    p.add_argument("--output_dir",        default="emb")
    p.add_argument("--device",            default="",
                   help="'cuda' / 'cpu' / '' (auto-detect)")
    return vars(p.parse_args())


if __name__ == "__main__":
    import sys
    if "ipykernel" in sys.modules:
        # ── Jupyter 模式：直接在这里修改参数 ──────────
        params = {
            "xlsx_path":         "majors_hierarchy.xlsx",
            "encode_model":      "hfl/chinese-roberta-wwm-ext",
            "encode_batch_size": 32,
            "encode_max_len":    512,
            "output_dir":        "emb",  # 新路径（相对路径）
            "device":            "",
        }
        encode(params)
    else:
        encode(parse_args())